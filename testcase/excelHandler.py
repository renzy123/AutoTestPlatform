#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Excel处理相关的操作的集合
# @author 任宗毅
# create_time : 2018/11/15

from openpyxl import load_workbook
from utils.consts import *
import os
from testcase.models import TestCase, CaseModule


class ExcelHandler:
    def __init__(self, file_name):
        """
        :param path:Excel的路径
        """
        self.file_name = str(file_name).rsplit(".", maxsplit=1)[0]
        self.workbook = load_workbook(filename=os.path.join(EXCELS_DIR, file_name))
        self.default_sheet_to_read = 2

    def modules_of_case(self):
        """生成相应的获取所有的Module和二级module"""

        class ModuleIndex:
            def __init__(self):
                self.start_index = 0
                self.end_index = 0
                self.name = ""

        # 从第二列获取所有的一级module
        # 从第三列获取所有的二级module
        # 默认获取的为读取第一个sheet
        sheet_list = self.workbook.sheetnames
        case_sheet = self.workbook[sheet_list[self.default_sheet_to_read - 1]]
        primary_modules = [_cell.value for _cell in case_sheet["B"]]
        primary_modules.pop(0)
        # 获取Primary_module所对应的相应的index列表
        current_index = 0
        primary_modules_indexes = []
        for _value in primary_modules:
            if _value:
                _module = ModuleIndex()
                _module.name = _value
                _module.start_index = current_index
                primary_modules_indexes.append(_module)
            primary_modules_indexes[-1].end_index = current_index
            current_index += 1
        # 获取secondary_module的index范围
        secondary_module = [_cell.value for _cell in case_sheet["C"]]
        secondary_module.pop(0)
        secondary_modules_indexes = []
        current_index = 0
        for _value in primary_modules:
            if _value:
                _module = ModuleIndex()
                _module.name = _value
                _module.start_index = current_index
                secondary_modules_indexes.append(_module)
            secondary_modules_indexes[-1].end_index = current_index
            current_index += 1

        # 将数据写入到module表中
        for _module in primary_modules:
            case_module = CaseModule()

        return {"primary_module": primary_modules_indexes, "secondary_module": secondary_modules_indexes}
