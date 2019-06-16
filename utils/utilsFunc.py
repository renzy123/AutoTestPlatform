#!/usr/bin/env python
# -*- coding: utf-8 -*-
# 工具方法模组
import datetime
import platform
import random

from openpyxl import load_workbook

from utils.consts import *


class BaseResult:
    """返回结果的默认格式"""

    # code=1:成功
    # code=2:失败

    def __init__(self):
        self.code = 1
        self.reason = None

    def toJSON(self):
        return {"code": self.code, "reason": self.reason}


def first_letter_of_chinese(string):
    def single_get_first(unicode1):
        str1 = unicode1.encode('gbk')
        try:
            ord(str1)
            return str1.decode('gbk')
        except:
            asc = str1[0] * 256 + str1[1] - 65536
            if -20319 <= asc <= -20284:
                return 'a'
            if -20283 <= asc <= -19776:
                return 'b'
            if -19775 <= asc <= -19219:
                return 'c'
            if -19218 <= asc <= -18711:
                return 'd'
            if -18710 <= asc <= -18527:
                return 'e'
            if -18526 <= asc <= -18240:
                return 'f'
            if -18239 <= asc <= -17923:
                return 'g'
            if -17922 <= asc <= -17418:
                return 'h'
            if -17417 <= asc <= -16475:
                return 'j'
            if -16474 <= asc <= -16213:
                return 'k'
            if -16212 <= asc <= -15641:
                return 'l'
            if -15640 <= asc <= -15166:
                return 'm'
            if -15165 <= asc <= -14923:
                return 'n'
            if -14922 <= asc <= -14915:
                return 'o'
            if -14914 <= asc <= -14631:
                return 'p'
            if -14630 <= asc <= -14150:
                return 'q'
            if -14149 <= asc <= -14091:
                return 'r'
            if -14090 <= asc <= -13119:
                return 's'
            if -13118 <= asc <= -12839:
                return 't'
            if -12838 <= asc <= -12557:
                return 'w'
            if -12556 <= asc <= -11848:
                return 'x'
            if -11847 <= asc <= -11056:
                return 'y'
            if -11055 <= asc <= -10247:
                return 'z'
            return ''

    if string is None:
        return None
    lst = list(string)
    charLst = []
    for l in lst:
        charLst.append(single_get_first(l))
    return ''.join(charLst)


def path_script_storage():
    """返回script的路径"""
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    storage_dir = os.path.join(BASE_DIR, "storage/scripts")
    return storage_dir


def base_path():
    """返回当前项目的根目录"""
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return BASE_DIR


def serialize_model(model):
    res = model.__dict__
    res.pop("_state")
    return res


def read_scripts(path):
    """
    读取测试用例的代码，用以进行展示
    """
    abs_path = os.path.join(SCRIPT_DIR, path)
    with open(abs_path, encoding="utf-8") as f:
        return f.readlines()


def current_os():
    """返回当前的操作系统"""
    _current_os = platform.platform()
    _current_os = _current_os.split("-", maxsplit=1)[0]
    if _current_os == "Darwin":
        return OS_MACOS
    if _current_os == "Windows":
        return OS_WINDOWS


def rename_file(file_name):
    """修改文件的命名，以符合操作系统规范"""
    inlegal_chars = ["\\", "/", ":", "*", "?", "''", "<", ">", "|"]

    for char in inlegal_chars:
        if char in file_name:
            file_name = str(file_name).replace(char, "_")
    return file_name


def gen_data_json(model, *args):
    """将模型文件和其他数据联合生成数据模型，并且返回json数据"""

    model_dict = model.__dict__
    for data in args:
        model_dict = {**model_dict, **data}
    return model_dict


def local_time_now():
    """格式化时间的显示"""
    time_format = "%Y-%m-%d %H:%M:%S"
    tz_utc_8 = datetime.timezone(datetime.timedelta(hours=8))
    now = datetime.datetime.now(tz=tz_utc_8)
    return now.strftime(time_format)


def gen_log_title(task_title):
    """生成日志的方法"""
    time_stamp = local_time_now()
    random_suffix = random.randint(0, 100)
    return rename_file(task_title + "_" + time_stamp + str(random_suffix) + ".log")


def read_excel():
    wb = load_workbook(r'C:\Users\renzy\Desktop\二人麻将\JJcd_九天项目群_二人麻将10期_功能需求case_导入用.xlsx')
    print(wb.sheetnames)
    sheet = wb.get_sheet_by_name("详细测试用例")
    print([cell.value for cell in sheet["C"]])


class ExcelHandler:
    def __init__(self, file_name):
        """
        :param path:Excel的路径
        """
        self.workbook = load_workbook(filename=file_name)
        self.default_sheet_to_read = 2

    def modules_of_case(self):
        """生成相应的获取所有的Module和二级module"""

        class ModuleIndex:
            def __init__(self):
                self.start_index = 0
                self.end_index = 0
                self.name = ""

        # 从第二列获取所有的一级module
        # 从第三列获取所有的二级module
        # 默认获取的为读取第一个sheet
        sheet_list = self.workbook.sheetnames
        case_sheet = self.workbook[sheet_list[self.default_sheet_to_read - 1]]
        primary_modules = [_cell.value for _cell in case_sheet["B"]]
        primary_modules.pop(0)
        # 获取Primary_module所对应的相应的index列表
        current_index = 0
        primary_modules_indexes = []
        for _value in primary_modules:
            if _value:
                _module = ModuleIndex()
                _module.name = _value
                _module.start_index = current_index
                primary_modules_indexes.append(_module)
            primary_modules_indexes[-1].end_index = current_index
            current_index += 1
        # 获取secondary_module的index范围
        secondary_module = [_cell.value for _cell in case_sheet["C"]]
        secondary_module.pop(0)
        secondary_modules_indexes = []
        current_index = 0
        for _value in primary_modules:
            if _value:
                _module = ModuleIndex()
                _module.name = _value
                _module.start_index = current_index
                secondary_modules_indexes.append(_module)
            secondary_modules_indexes[-1].end_index = current_index
            current_index += 1
        return {"primary_module": primary_modules_indexes, "secondary_module": secondary_modules_indexes}


if __name__ == '__main__':
    excel_handler = ExcelHandler(r'C:\Users\renzy\Desktop\二人麻将\JJcd_九天项目群_二人麻将10期_功能需求case_导入用.xlsx')
    print(excel_handler.modules_of_case())
